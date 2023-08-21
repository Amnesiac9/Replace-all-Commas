import os
from openpyxl import load_workbook
from openpyxl import Workbook
import xlrd
import xlwt
import pandas as pd
import numpy as np


# Define the directory where your .xlsx files are located
#directory_path = '/path/to/your/directory'

decimal_places = 5

directory_path = input("enter the full directory path: ")

# Define the text you want to find and replace
find_text = ','
replace_text = ''

# Function to format 'Gallons' column
def format_gallons(value):
    if isinstance(value, (float, int)):
        return round(value, 5)
    else:
        return value
    
    
# Specify the column to iterate over
column_letter = 'O'

# Define a function to round and format a number
# def format_number(value):
#     if isinstance(value, (int, float)):
#         return f'{value:.2f}'
#     return value  # Return non-numeric values as they are


def is_numeric(value):
    try:
        numeric_value = float(value)
        if numeric_value.is_integer():
            return True
        else:
            return True
    except (ValueError, TypeError):
        return False


for filename in os.listdir(directory_path):
    if filename.endswith('.xls'):
        file_path = os.path.join(directory_path, filename)

        
        # Read the .xls file into a pandas DataFrame
        df = pd.read_excel(file_path)
        
        # Set headers
        df.columns = ['SKU', 'Name', 'Blank', 'Blank2', 'Tax Class', 'Size', 'On Hand Cases', 'On Hand Bottles', 'Open Order Cases', 'Open Order Bottles', 'Available Cases', 'Available Bottles', 'Cost/Case', 'On-Hand Value', 'Gallons']

        df.dropna(how='all', inplace=True)
        df = df.fillna('')
        # Perform the find and replace operation on all DataFrame columns
        df = df.applymap(lambda x: str(x).replace(find_text, replace_text))

        # Save the DataFrame to an .xlsx file
        xlsx_filename = filename.replace('.xls', '.xlsx')
        xlsx_file_path = os.path.join(directory_path, xlsx_filename)
       # df.to_excel(xlsx_file_path, index=False)
        
        # Create a new Excel workbook and add the DataFrame to it
        wb = Workbook()
        ws = wb.active
        ws.title = 'IN0110_32.RPT'
        ws.append(list(df.columns))
        # Append DataFrame rows to the worksheet
        for index, row in df.iterrows():
            ws.append(list(row))
            
        # Iterate over the cells in the specified column
        for cell in ws[column_letter]:
            # Check if it's a numeric value
            if is_numeric(cell.value):
                cell.data_type = 'n'
                cell.number_format = '0.00000'
            
            
        try:
            wb.save(xlsx_file_path)
        except PermissionError:
            wb.save(xlsx_file_path+'_2.xlsx')
    

print("Commas replaced in all .xls files.")