import os
from openpyxl import Workbook
import pandas as pd


# Define the directory where your .xlsx files are located
#directory_path = '/path/to/your/directory'

DECIMAL_PLACES = 5

directory_path = current_directory = os.getcwd() #input("enter the full directory path: ")


# Define the text you want to find and replace
find_text = ','
replace_text = ''

# Function to format 'Gallons' column
def format_gallons(value):
    if isinstance(value, (float, int)):
        return round(value, DECIMAL_PLACES)
    else:
        return value
    
    


# Define a function to round and format a number
# def format_number(value):
#     if isinstance(value, (int, float)):
#         return f'{value:.2f}'
#     return value  # Return non-numeric values as they are


def is_numeric(value):
    try:
        numeric_value = float(value)
        if numeric_value.is_integer():
            return True
        else:
            return True
    except (ValueError, TypeError):
        return False


for filename in os.listdir(directory_path):
    if filename.endswith('.xls'):
        file_path = os.path.join(directory_path, filename)

        
        # Read the .xls file into a pandas DataFrame
        df = pd.read_excel(file_path)
        
        # Set headers
        try: 
            df.columns = ['SKU', 'Name', 'Blank', 'Blank2', 'Tax Class', 'Size', 'On Hand Cases', 'On Hand Bottles', 'Open Order Cases', 'Open Order Bottles', 'Available Cases', 'Available Bottles', 'Cost/Case', 'On-Hand Value', 'Gallons']
        except:
            df.columns = ['SKU', 'Name', 'Blank', 'Blank2', 'Tax Class', 'Size', 'On Hand Cases', 'On Hand Bottles', 'Open Order Cases', 'Open Order Bottles', 'Available Cases', 'Available Bottles']
            print(f"No data in {filename}, implimenting shorter columns")
    
        
        # print("df.columns:", df.columns)
        
        # Grab the name of the file to compare to internal contents
        filenameSplit = filename.split('.')[0]
        try:
            reportName = df['SKU'][23].split(' ')[1][:5] #grab the first 5 characters of the 1 index of the report name line
            if str.lower(filenameSplit) != str.lower(reportName):
                print('Warning! ReportName:  does not match File Name!')
                print(f'filename: {filenameSplit} | reportname: {reportName}')
        except:
            print('Error grabbing file name from cell A25. Filenames not compared.')
            print(f'filename: {filenameSplit} | reportname: {reportName}')
        
            
        
        
    

        df.dropna(how='all', inplace=True)
        df = df.fillna('')
        # Perform the find and replace operation on all DataFrame columns
        #df = df.applymap(lambda x: str(x).replace(find_text, replace_text))
        # df = df.applymap(lambda x: str(x).replace(find_text, replace_text))
        df = df.apply(lambda x: x.map(lambda val: str(val).replace(find_text, replace_text)))
        

        
        # Get the sum of the available cases column for later
        availableCasesSum = 0
        for value in df['On Hand Cases']:
            try:
                num = int(value)
                availableCasesSum += num
            except ValueError:
                pass
      
        xlsx_filename = filename.replace('.xls', '.xlsx')
        xlsx_file_path = os.path.join(directory_path, xlsx_filename)
       # df.to_excel(xlsx_file_path, index=False)
       
       # variable to track worksheet sum of available cases
        availableCasesSumWB = 0
        
        # Create a new Excel workbook and add the DataFrame to it
        wb = Workbook()
        ws = wb.active
        ws.title = 'IN0110_32.RPT'
        ws.append(list(df.columns))
        # Append DataFrame rows to the worksheet
        for index, row in df.iterrows():
            ws.append(list(row))
            
        # Specify the column to iterate over for gallons
        COLUMN_LETTER = 'O'
        # Iterate over the cells in the specified column
        for cell in ws[COLUMN_LETTER]:
            # Check if it's a numeric value
            if is_numeric(cell.value):
                cell.data_type = 'n'
                cell.number_format = '0.00000'
                
        for row in ws.iter_rows(min_row=9):
            for cell in row:
                if is_numeric(cell.value):
                    cell.data_type = 'n'
                    
        for cell in ws['G']:
            try:
                num = int(cell.value)
                availableCasesSumWB += num
            except ValueError:
                pass
        
        #Add the headers to row 11
        for cell in ws[11]:
            col_num = cell.column - 1
            try:
                cell.value = ws[1][col_num].value # pull the headers and put in row 11
            except:
                print(f"col_num out of bounds on row 11, Col_num: {col_num}, cell column: {cell.column}")
        
        ws.delete_rows(1,1) # Delete the hold headers at the top
        ws.insert_rows(10, 1) # Add blank row at index 10 to seperate header info from the table
            
        if availableCasesSum != availableCasesSumWB:
            print(f'WARNING! Sums don\'t match! Original sum: {availableCasesSum}, wb sum: {availableCasesSumWB}')
            
        try:
            wb.save(xlsx_file_path)
        except PermissionError:
            wb.save(xlsx_file_path+'_2.xlsx')
    

print("Commas replaced in all .xls files. v1.1.0")